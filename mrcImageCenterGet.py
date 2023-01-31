import subprocess
import openpyxl
import os.path
import sys


exel_file = sys.argv[1]
exel_data = openpyxl.load_workbook(exel_file)
exel_sheet = sys.argv[2]
raw_data = sys.argv[3]
x_wide = sys.argv[4]
y_wide = sys.argv[5]
z_wide = sys.argv[6]

#エクセルの行数取得
sheet = exel_data[f'{exel_sheet}']
index = sheet.max_row

num =[]
cx = []
cy = []
cz = []

for i in range(0,index-1,1):
    num.insert(i,sheet.cell(row=i+2,column=1).value)
    #x軸の幅の開始地点(2列目B)
    cx.insert(i,sheet.cell(row=i+2,column=2).value)
    #y軸の幅の開始地点(3列目C)
    cy.insert(i,sheet.cell(row=i+2,column=3).value)
    #z軸の幅の開始地点(4列目D)
    cz.insert(i,sheet.cell(row=i+2,column=4).value)


for i in range(0,index-1,1):
    subprocess.run(['mrcImageCenterGet', '-i', f'{raw_data}','-o', f'bilateral_sub{num[i]}.mrc', '-Cx', f'{cx[i]}', '-Cy', f'{cy[i]}', '-Cz', f'{cz[i]}', '-Nx', f'{x_wide}', '-Ny', f'{y_wide}', '-Nz', f'{z_wide}'])
